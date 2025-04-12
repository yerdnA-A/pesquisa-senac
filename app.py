from flask import Flask, request, redirect, url_for, render_template
from openpyxl import Workbook, load_workbook
from openpyxl.utils import *
import os
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# Função para obter as credenciais de acesso
def obter_credenciais():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    
    # Se não houver credenciais válidas, peça ao usuário para fazer login.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Salve as credenciais para a próxima vez que o programa rodar
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

# Função para acessar a planilha
def acessar_planilha():
    creds = obter_credenciais()
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    return sheet


app = Flask(__name__)
EXCEL_FILE = 'respostas.xlsx'

# Garante que o arquivo Excel existe
def inicializa_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Respostas'
        ws.append([
            'Nome', 'Idade', 'Curso/Função', 'Possui Computador',
            'Possui Celular', 'Possui Antivírus', 'Escaneamento',
            'Atualizações', 'Reutiliza Senhas', 'Senhas com Símbolos',
            'Verifica Domínio', 'IP'
        ])
        wb.save(EXCEL_FILE)

@app.route('/enviar', methods=['POST'])
def enviar():
    data = request.form
    ip_usuario = request.remote_addr


    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        data.get('nome'),
        data.get('idade'),
        data.get('funcao'),
        data.get('computador'),
        data.get('celular'),
        data.get('antivirus'),
        data.get('scan'),
        data.get('atualiza'),
        data.get('reutiliza'),
        data.get('simbolos'),
        data.get('dominio'),
        ip_usuario
    ])

    dados_google = {
        'nome': data.get('nome'),
        'idade': data.get('idade'),
        'funcao': data.get('funcao'),
        'computador': data.get('computador'),
        'celular': data.get('celular'),
        'antivirus': data.get('antivirus'),
        'scan': data.get('scan'),
        'atualiza': data.get('atualiza'),
        'reutiliza': data.get('reutiliza'),
        'simbolos': data.get('simbolos'),
        'dominio': data.get('dominio'),
        'ip_usuario': ip_usuario
    }
    adicionar_resposta(dados_google)


    wb.save(EXCEL_FILE)
    return redirect(url_for('obrigado'))

@app.route('/obrigado')
def obrigado():
    return render_template('obrigado.html')

@app.route('/')
def index():
    return render_template('index.html')

def adicionar_resposta(dados):
    sheet = acessar_planilha()
    SPREADSHEET_ID = '14QdRiGB4a_ebYA0casBXxu2hSNAb_xUHe2ZmVA91nuY'  # Substitua pela ID da sua planilha
    RANGE_NAME = 'Respostas!A1'
    valores = [
        [
            dados['nome'], dados['idade'], dados['funcao'], dados['computador'],
            dados['celular'], dados['antivirus'], dados['scan'], dados['atualiza'],
            dados['reutiliza'], dados['simbolos'], dados['dominio'], dados['ip_usuario']
        ]
    ]
    body = {
        'values': valores
    }
    result = sheet.values().append(
        spreadsheetId=SPREADSHEET_ID, range=RANGE_NAME,
        valueInputOption="RAW", body=body).execute()
    return result


if __name__ == '__main__':
    inicializa_excel()
    app.run(debug=True)